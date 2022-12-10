from PyPDF2 import PdfFileReader
import zipfile
import csv
from openpyxl import load_workbook

with zipfile.ZipFile('data/spam.zip', 'w') as myzip:
    myzip.write('data/some.csv')
    myzip.write('data/simple_demo.pdf')
    myzip.write('data/test_with_files.xlsx')  # работа с архивированием в zip
    a = myzip.namelist()
    print(a)
    print(a==['data/some.csv', 'data/simple_demo.pdf', 'data/test_with_files.xlsx'])

with open('data/some.csv', 'w', newline='') as f:
    writer = csv.writer(f, delimiter=',')
    writer.writerow(['how are you', 'Anna'])  # работа с csv


with open('data/some.csv', newline='') as f:
    rider = csv.reader(f)
    for row in rider:
        print(row)
        print(row==['how are you', 'Anna'])

pdf_document = "data/simple_demo.pdf"
with open(pdf_document, "rb") as filehandle:
    pdf = PdfFileReader(filehandle)
    info = pdf.getDocumentInfo()
    pages = pdf.getNumPages()
    print(f'\n {info}')
    print(f'\nnumber of pages: {pages}\n')
    page1 = pdf.getPage(0)
    text=page1.extractText()
    print(page1.extractText())
    print(text=='Welcome to Python!')# работа с pdf

book = load_workbook('data/test_with_files.xlsx')
sheet = book.active
ass=sheet.cell(row=3, column=2).value
print(ass)
print(ass=='Болгария')# работа с xlsx

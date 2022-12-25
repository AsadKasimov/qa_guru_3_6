import re
from PyPDF2 import PdfFileReader
import zipfile
import csv
from openpyxl import load_workbook


# тут только тесты с файлами

def test_creat_zip():
    with zipfile.ZipFile('data/spam.zip', 'w') as myzip:
        myzip.write('data/some.csv')
        myzip.write('data/simple_demo.pdf')
        myzip.write('data/test_with_files.xlsx')


def test_zip():
    with zipfile.ZipFile('data/spam.zip', 'r') as myzip:
        assert len(myzip.namelist()) == 3  # проверка сколько файлов входтит в файл
        assert 'data/simple_demo.pdf' in myzip
        assert 'data/some.csv' in myzip
        assert 'data/test_with_files.xlsx' in myzip
    assert 'spam.zip' in 'data'

def test_csv():
    # тут читаем csv файлы
    with open('data/some.csv', newline='') as f:
        for row in csv.reader(f):
            assert 'California' in row
            assert ['California', 'Sacramento', 'Los Angeles', '39538223'] == row
            assert len(row) == 4


# тут раболта с pdf
def test_pdf():
    pdf_document = "data/simple_demo.pdf"
    with open(pdf_document, "rb") as filehandle:
        pdf = PdfFileReader(filehandle)
        page1 = pdf.getPage(0)
        text = page1.extractText()
        # тесты
        print(re.findall(r'^\w+', text, flags=re.IGNORECASE) == ['Welcome'])
        text_from = re.findall(r'\w+', text, flags=re.IGNORECASE)
        assert 'Welcome' in text_from
        assert 'to' in text_from
        assert 'Python' in text_from
        assert len(text_from) == 3

def test_ex():
    # тут читаем exel
    book = load_workbook('data/test_with_files.xlsx')
    sheet = book.active
    ass = sheet.cell(row=3, column=3).value
    assert sheet.max_row == 11
    assert sheet.max_column == 8
    assert ass == 25000
